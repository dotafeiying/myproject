#-*- coding:utf-8 -*-
import MySQLdb,codecs,xlwt
from io import StringIO,BytesIO
import pandas as pd
import openpyxl,gc
from impala.dbapi import connect
from django.http import StreamingHttpResponse,HttpResponse

def csv_stream_response_generator(export_sql):
    db = MySQLdb.connect("10.39.211.198", "root", "password", "busycell", charset='utf8')
    chunk_size = 30000
    offset = 0
    yield codecs.BOM_UTF8
    while True:
        isHeader = False
        print(offset)
        if offset == 0:
            isHeader = True
        # sql = "SELECT id,enbid,cellid,cellname,freqID,lng,lat,scene,indoor,结果 FROM busycell limit %d offset %d" % (chunk_size, offset)
        # sql= "select * from busycell where (enbid like '%{0}%' or cellname like '%{0}%') " \
        #             "and unix_timestamp(finish_time)>unix_timestamp('{3}') and unix_timestamp(finish_time)<unix_timestamp('{4}') " \
        #             "and (city in ({5}) or concat(enbid,'_',cellid) in ({6})) " \
        #             "order by finish_time limit {2} offset {1}".format(
        #     name, offset, chunk_size, dateStart, dateEnd, str(citys)[1:-1], str(cells)[1:-1])
        # sql=export_sql
        sql=export_sql + " limit {1} offset {0}".format(offset, chunk_size)
        print(sql)
        df = pd.read_sql(sql, db)
        f = StringIO()
        df.to_csv(f, index=False, header=isHeader, encoding="utf_8_sig")
        yield f.getvalue()
        offset += chunk_size
        if df.shape[0] < chunk_size:
            break

def csv_stream_response_generator1(**kwargs):
    db = MySQLdb.connect("10.39.211.198", "root", "password", "busycell", charset='utf8')
    chunk_size = 30000
    offset = 0
    yield codecs.BOM_UTF8
    while True:
        isHeader = False
        print(offset)
        if offset == 0:
            isHeader = True
        name = kwargs.get('name')
        citys = kwargs.get('citys')
        cells = kwargs.get('cells')
        dateStart = kwargs.get('dateStart')
        dateEnd = kwargs.get('dateEnd')
        print(citys)
        print(cells)
        # sql = "SELECT id,enbid,cellid,cellname,freqID,lng,lat,scene,indoor,结果 FROM busycell limit %d offset %d" % (chunk_size, offset)
        sql= "select * from busycell where (enbid like '%{0}%' or cellname like '%{0}%') " \
                    "and unix_timestamp(finish_time)>unix_timestamp('{3}') and unix_timestamp(finish_time)<unix_timestamp('{4}') " \
                    "and (city in ({5}) or concat(enbid,'_',cellid) in ({6})) " \
                    "order by finish_time limit {2} offset {1}".format(
            name, offset, chunk_size, dateStart, dateEnd, str(citys)[1:-1], str(cells)[1:-1])
        df = pd.read_sql(sql, db)
        f = StringIO()
        df.to_csv(f, index=False, header=isHeader, encoding="utf_8_sig")
        yield f.getvalue()
        offset += chunk_size
        if df.shape[0] < chunk_size:
            break

import re
import numbers
import six
import tablib
from django.utils.encoding import smart_str
def nullify(cell):
  return cell if cell is not None else "NULL"
ILLEGAL_CHARS = r'[\000-\010]|[\013-\014]|[\016-\037]'

def encode_row(row, encoding=None, make_excel_links=False):
  encoded_row = []
  for cell in row:
    if isinstance(cell, six.string_types):
      cell = re.sub(ILLEGAL_CHARS, '?', cell)
      if make_excel_links:
        cell = re.compile('(https?://.+)', re.IGNORECASE).sub(r'=HYPERLINK("\1")', cell)
    cell = nullify(cell)
    if not isinstance(cell, numbers.Number):
      cell = smart_str(cell, encoding or 'utf-8', strings_only=True, errors='replace')
    encoded_row.append(cell)
  return encoded_row

class XlsWrapper():
  def __init__(self, xls):
    self.xls = xls


def xls_dataset(workbook):
  output = BytesIO()
  workbook.save(output)
  output.seek(0)
  return XlsWrapper(output.read())

class HS2DataAdapter:
  def __init__(self, conn):

    self.conn = conn
    self.offset=0
    self.sql = None
    self.fetch_size = 1000

    self.first_fetched = True
    self.headers = None
    self.num_cols = None
    self.row_counter = 1
    self.bytes_counter = 0
    self.is_truncated = False
    self.has_more = True

  def __iter__(self):
    return self

  def __next__(self):
      self.sql = "SELECT enbid,cellid,cellname,freqID,lng,lat,scene,indoor,result FROM lte_busy_cell_history order by finish_time limit %d offset %d" % (
          10000, self.offset)
      results = pd.read_sql(self.sql,self.conn)
      shape=results.shape[0]
      self.offset+= 10000
      print(self.offset)
      print('shape:',shape)
      if self.first_fetched:
          self.first_fetched = False
          headers = results.columns.tolist()
          self.headers = headers
      if self.has_more:
          rows = results.values.tolist()
          data=rows
          return self.headers,data
      if shape < 10000:
          self.has_more=False
          raise StopIteration()


def dataset(headers, data, encoding=None):
  """
  dataset(headers, data) -> Dataset object
  Return a dataset object for a csv or excel document.
  """
  dataset = tablib.Dataset()
  if headers:
    dataset.headers = encode_row(headers, encoding)
  for row in data:
    dataset.append(encode_row(row, encoding))
  return dataset

# def make_response(generator, format, name, encoding=None, user_agent=None):
#   """
#   @param data An iterator of rows, where every row is a list of strings
#   @param format Either "csv" or "xls"
#   @param name Base name for output file
#   @param encoding Unicode encoding for data
#   """
#   if format == 'csv':
#     content_type = 'application/csv'
#     resp = StreamingHttpResponse(generator, content_type=content_type)
#     try:
#       del resp['Content-Length']
#     except KeyError:
#       pass
#   elif format == 'xls':
#     format = 'xlsx'
#     content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     resp = HttpResponse(next(generator), content_type=content_type)
#
#   elif format == 'json':
#     content_type = 'application/json'
#     resp = HttpResponse(generator, content_type=content_type)
#   else:
#     raise Exception("Unknown format: %s" % format)
#
#   try:
#     name = name.encode('ascii')
#     resp['Content-Disposition'] = 'attachment; filename="%s.%s"' % (name, format)
#   except UnicodeEncodeError:
#     name = urlquote(name)
#     if user_agent is not None and 'Firefox' in user_agent:
#       # Preserving non-ASCII filename. See RFC https://tools.ietf.org/html/rfc6266#appendix-D, only FF works
#       resp['Content-Disposition'] = 'attachment; filename*="%s.%s"' % (name, format)
#     else:
#       resp['Content-Disposition'] = 'attachment; filename="%s.%s"' % (name, format)
#
#   return resp

def create_generator(content_generator, format, encoding=None):
  if format == 'csv':
    show_headers = True
    for headers, data in content_generator:
      yield dataset(show_headers and headers or None, data, encoding).csv
      show_headers = False
  elif format == 'xls':
    workbook = openpyxl.Workbook(write_only=True)
    worksheet = workbook.create_sheet()
    row_ctr = 0

    for _headers, _data in content_generator:
      # Write headers to workbook once
      if _headers and row_ctr == 0:
        worksheet.append(encode_row(_headers, encoding))
        row_ctr += 1

      # Write row data to workbook
      for row in _data:
        worksheet.append(encode_row(row, encoding, make_excel_links=True))
        row_ctr += 1
    print(row_ctr)
    yield xls_dataset(workbook).xls
    gc.collect()
  else:
    raise Exception("Unknown format: %s" % format)

def excel_stream_response_generator():
    # db = MySQLdb.connect("10.39.211.198", "root", "password", "test", charset='utf8')
    db = connect(host='133.21.254.164', port=21050, database='hub_yuan')
    chunk_size = 20000
    offset = 0
    # content_generator = HS2DataAdapter(db)
    # generator = create_generator(content_generator, 'xls')
    # content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    # resp = StreamingHttpResponse(generator, content_type=content_type)
    # resp['Content-Disposition'] = 'attachment; filename="%s.%s"' % ('test', 'xls')
    # return resp

    sql = "SELECT index,enbid,cellid,cellname,freqID,lng,lat,scene,indoor,result FROM lte_busy_cell_history"
    data=pd.read_sql(sql,db,chunksize=20000)
    for df in data:
        rows=df.values.tolist()
        # print(df.dtypes)
        yield tablib.Dataset(*rows).xlsx
        gc.collect()
    # print(type(data))
    # return ([tablib.Dataset(*df.values.tolist()).xls for df in data])

    # while True:
    #     sql = "SELECT enbid,cellid,cellname,freqID,lng,lat,scene,indoor,result FROM lte_busy_cell_history order by finish_time limit %d offset %d" % (
    #     chunk_size, offset)
    #     data = pd.read_sql(sql, db)
    #     # f = BytesIO()
    #     # rows = data.values.tolist()
    #     # headers = data.columns.tolist()
    #     # nrows = len(rows)
    #     # print(rows)
    #     output = BytesIO()
    #     writer = pd.ExcelWriter(output, engine='xlsxwriter')
    #     data.to_excel(writer)
    #     # for row in rows:
    #     #     table.append(row)
    #     # workbook.save(f)
    #     yield output.getvalue()
    #     writer.save()
    #     # yield xls_dataset(workbook).xls
    #     gc.collect()
    #     offset += chunk_size
    #     if data.shape[0] < chunk_size:
    #         break

def xlsx_stream_response_generator(file_name,chunk_size=512):
    # db = MySQLdb.connect("10.39.211.198", "root", "password", "test", charset='utf8')
    with open(file_name) as f:
        while True:
            c = f.read(chunk_size)
            if c:
                yield c
            else:
                break


import os,django,sys
BASE_DIR =os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.append(BASE_DIR)


import datetime
import xlwt


# def BulidNewExcel(download_url,data):
#     style0 = xlwt.easyxf('font: name Times New Roman, color-index red, bold on')
#     style1 = xlwt.easyxf(num_format_str='D-MMM-YY')
#
#     db = MySQLdb.connect("10.39.211.198", "root", "password", "test", charset='utf8')
#     sql="select * from busycell where enbid like '%{0}%' or cellname like '%{0}%' limit {1},{2}".format(name,offset,limit)
#     sql = "SELECT id,enbid,cellid,cellname,freqID,lng,lat,scene,indoor,结果 FROM busycell limit %d offset %d" % (
#     chunk_size, offset)
#     data = pd.read_sql(sql, db)
#
#     wb = xlwt.Workbook()
#     ws = wb.add_sheet('Sheet',cell_overwrite_ok=True)
#     for i in range(len(field_verbose_name_list)):
#         ws.write(0,i,field_verbose_name_list[i],style0)
#     timestr=datetime.datetime.now().strftime("%Y%m%d%H%M%S")
#     wb.save(download_url+'New-'+timestr+'.xls')
#     return timestr